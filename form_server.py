from flask import Flask, request, render_template, jsonify, send_from_directory
from openpyxl import load_workbook, Workbook
import os

# --- Configuration ---
# Excel file settings
EXCEL_FILE = 'form_data.xlsx'

# Web server settings
# Host '0.0.0.0' makes the server accessible on your local network
HOST = '0.0.0.0' 
PORT = 5666 

# --- Flask Web Server ---
app = Flask(__name__, static_folder='.', template_folder='.')

@app.route('/')
def form():
    """Serves the main HTML form."""
    return render_template('index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    """Serves static files like CSS and JS."""
    return send_from_directory('.', filename)

@app.route('/submit', methods=['POST'])
def submit():
    """Handles form submission and saves data to Excel."""
    try:
        data = request.json
        print("Received data:", data)

        # Define the order of columns
        headers = [
            'department', 'activityTopic', 'activityType', 'strategicGoalLevel1',
            'strategicGoalLevel2', 'presenterCategory', 'activityDate', 'presenterName',
            'attendanceResponsible', 'targetAudience', 'attendeeCount', 'activityDuration',
            'contentLocation'
        ]

        # Prepare the row with data in the correct order
        row_data = [data.get(header, '') for header in headers]

        # Check if the Excel file exists, create it if it doesn't
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Form Submissions"
            # Create more readable headers for the Excel file
            readable_headers = [
                'Department', 'Activity Topic', 'Activity Type', 'Strategic Goal L1',
                'Strategic Goal L2', 'Presenter Category', 'Activity Date', 'Presenter Name',
                'Attendance Responsible', 'Target Audience', 'Attendee Count', 
                'Activity Duration (Hours)', 'Content Location'
            ]
            ws.append(readable_headers)
            wb.save(EXCEL_FILE)

        # Load the workbook and append the new data
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(row_data)
        wb.save(EXCEL_FILE)

        return jsonify({'message': 'Data saved successfully!'}), 200
    except Exception as e:
        print(f"Error saving data: {e}")
        return jsonify({'message': f'An error occurred: {e}'}), 500

# --- Main Execution ---
if __name__ == '__main__':
    print("Starting the form server...")
    print(f"To fill out the form, please open a web browser and go to:")
    print(f"http://127.0.0.1:{PORT} (on this computer)")
    print(f"Or use your local network IP address from other devices on the same network.")
    print("Press CTRL+C to stop the server.")
    app.run(host=HOST, port=PORT)