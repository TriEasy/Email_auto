import smtplib
import schedule
import time
import threading
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Flask, request, render_template, jsonify, send_from_directory
from openpyxl import load_workbook, Workbook
import os

# --- Configuration ---
# Email settings
SMTP_SERVER = 'smtp.gmail.com'  # Replace with your SMTP server
SMTP_PORT = 587
EMAIL_SENDER = 'fahudeah@gmail.com'  # Replace with your email
EMAIL_PASSWORD = 'ncna qxse yxkt xwzf'      # Replace with your email app password
EMAIL_SUBJECT = 'Monthly Knowledge Activity Form'
EMPLOYEE_LIST_FILE = 'employees.txt'
FORM_FILE_PATH = 'index.html'

# Excel file settings
EXCEL_FILE = 'form_data.xlsx'

# Web server settings
HOST = '0.0.0.0'
PORT = 5000

# --- Flask Web Server ---
app = Flask(__name__, static_folder='.', template_folder='.')

@app.route('/')
def form():
    """Serves the main HTML form."""
    return render_template(FORM_FILE_PATH)

@app.route('/<path:filename>')
def serve_static(filename):
    """Serves static files like CSS and JS."""
    return send_from_directory('.', filename)

@app.route('/submit', methods=['POST'])
def submit():
    """Handles form submission and saves data to Excel."""
    try:
        data = request.json
        print("Received data:", data)

        # Define the order of columns
        headers = [
            'department', 'activityTopic', 'activityType', 'strategicGoalLevel1',
            'strategicGoalLevel2', 'presenterCategory', 'activityDate', 'presenterName',
            'attendanceResponsible', 'targetAudience', 'attendeeCount', 'activityDuration',
            'contentLocation'
        ]

        # Prepare the row with data in the correct order
        row_data = [data.get(header, '') for header in headers]

        # Check if the Excel file exists
        if not os.path.exists(EXCEL_FILE):
            # Create a new workbook and add the header row
            wb = Workbook()
            ws = wb.active
            ws.title = "Form Submissions"
            # Make headers more readable
            readable_headers = [h.replace('_', ' ').title() for h in headers]
            ws.append(readable_headers)
            wb.save(EXCEL_FILE)

        # Load the workbook and append the new data
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(row_data)
        wb.save(EXCEL_FILE)

        return jsonify({'message': 'Data saved successfully!'}), 200
    except Exception as e:
        print(f"Error saving data: {e}")
        return jsonify({'message': f'An error occurred: {e}'}), 500

def run_flask_app():
    """Runs the Flask web server."""
    print(f"Starting web server at http://{HOST}:{PORT}")
    app.run(host=HOST, port=PORT)

# --- Email Sending Logic ---
def get_employee_emails():
    """Reads employee emails from a text file."""
    try:
        with open(EMPLOYEE_LIST_FILE, 'r') as f:
            return [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        print(f"Error: The file '{EMPLOYEE_LIST_FILE}' was not found.")
        return []

def send_email(receiver_email, form_html):
    """Sends an email with the HTML form to a single recipient."""
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = EMAIL_SUBJECT
        msg['From'] = EMAIL_SENDER
        msg['To'] = receiver_email

        # Attach the HTML content
        part = MIMEText(form_html, 'html')
        msg.attach(part)

        # Send the email
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, receiver_email, msg.as_string())
            print(f"Email sent successfully to {receiver_email}")

    except Exception as e:
        print(f"Failed to send email to {receiver_email}. Error: {e}")

def schedule_monthly_email():
    """Reads the form and sends it to all employees."""
    print("Sending monthly form emails...")
    try:
        with open(FORM_FILE_PATH, 'r', encoding='utf-8') as f:
            form_content = f.read()
    except FileNotFoundError:
        print(f"Error: The form file '{FORM_FILE_PATH}' was not found.")
        return

    employees = get_employee_emails()
    if not employees:
        print("No employee emails found. Aborting email schedule.")
        return

    for email in employees:
        send_email(email, form_content)

# --- Scheduler ---
def run_scheduler():
    """Sets up and runs the email sending schedule."""
    # Schedule the job to run on the 1st day of every month at 9:00 AM
    schedule.every().month.on(1, "09:00").do(schedule_monthly_email)
    
    # For testing: run every minute
    # schedule.every(1).minutes.do(schedule_monthly_email)

    print("Scheduler started. Emails will be sent on the 1st of each month at 9:00 AM.")
    while True:
        schedule.run_pending()
        time.sleep(1)

# --- Main Execution ---
if __name__ == '__main__':
    # Start the Flask app in a separate thread
    flask_thread = threading.Thread(target=run_flask_app)
    flask_thread.daemon = True
    flask_thread.start()

    # Start the scheduler
    run_scheduler()